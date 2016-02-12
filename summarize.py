import os
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl import Workbook
import pandas as pd
import numpy as np
import xlrd
import copy

root_loc = "/home/varunwachaspati/LVPEI/"

#FOR DEBUGGING AGE_COUNT_SUMMARY
def compare(uid_dict):
    os.chdir(root_loc+"result")
    a = list(pd.read_csv("blvmn_data_below_10.csv")["uid"].unique())
    a+=list(pd.read_csv("blvmn_data_above_10.csv")["uid"].unique())
    a+=list(pd.read_csv("kvc_data_above_10.csv")["uid"].unique())
    a+=list(pd.read_csv("mdcc_data_above_10.csv")["uid"].unique())
    a+=list(pd.read_csv("kvc_data_below_10.csv")["uid"].unique())
    a+=list(pd.read_csv("mdcc_data_below_10.csv")["uid"].unique())
    b = []
    for x in a:
        if x not in uid_dict:
            print x
            b.append(x)
            print "been here"
    print len(set(a)),len(uid_dict),len(b), "Hello in Compare"

def age_count_summary():
    os.chdir(root_loc)
    age_count_dict = {}
    workbook = load_workbook(filename = "LVPEI_Refractive_Error_Data_Colorized.xlsx")
    for ws in workbook.worksheets:
        uid_dict = {}
        row_list = ws.rows[1:]
        for row in row_list:
            if row[0].value not in uid_dict:
                uid_dict[row[0].value] = (row[2].value,1)
            else:
                uid_dict[row[0].value] = list(uid_dict[row[0].value])
                uid_dict[row[0].value][1] += 1
                uid_dict[row[0].value] = tuple(uid_dict[row[0].value])

    #compare(uid_dict)

        for uid, visit_tuple in uid_dict.iteritems():
            if visit_tuple[0] in age_count_dict:
                if visit_tuple[1] < 2:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][0] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])
                elif visit_tuple[1] > 5:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][2] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])
                else:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][1] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])
            else:
                age_count_dict[visit_tuple[0]] = (0,0,0)
                if visit_tuple[1] < 2:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][0] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])
                elif visit_tuple[1] > 5:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][2] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])
                else:
                    age_count_dict[visit_tuple[0]] = list(age_count_dict[visit_tuple[0]])
                    age_count_dict[visit_tuple[0]][1] += 1
                    age_count_dict[visit_tuple[0]] = tuple(age_count_dict[visit_tuple[0]])

    return age_count_dict
    
def age_summary_write(age_count_dict):
    workbook = load_workbook(filename = "LVPEI_Refractive_Error_Data_Colorized.xlsx")
    new_ws = workbook.create_sheet()
    new_ws.title = "Patient Visit Count"
    new_ws["A1"].value = "Age"
    new_ws["B1"].value = "Visits < 2"
    new_ws["C1"].value = "Visits 2-5"
    new_ws["D1"].value = "Visits > 5"
    astr = 2
    for age, age_count_tuple in age_count_dict.iteritems():
            new_ws["A"+str(astr)].value = age
            new_ws["B"+str(astr)].value = age_count_tuple[0]
            new_ws["C"+str(astr)].value = age_count_tuple[1]
            new_ws["D"+str(astr)].value = age_count_tuple[2]
            astr += 1
    workbook.save("LVPEI_Refractive_Error_Data_Colorized_Summarized.xlsx")

def analysis():
    #VERY MESSY DATA STRUCTURE, REFACTOR !
    category_count = {'simple_myopia_high':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_myopia_moderate':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_myopia_low':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_hyperopia_high':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_hyperopia_moderate':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_hyperopia_low':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_myopic_astigmatism':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'simple_hypermetropic_astigmatism':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'compound_myopic_astigmatism':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'compound_hypermetropic_astigmatism':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'mixed_astigmatism_postive_sph':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'mixed_astigmatism_negative_sph':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'normal':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'undefined':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}},'Insufficient Data':{'os':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0},'od':{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0,25:0}}}
    zero_category_count = copy.deepcopy(category_count)
    first_category_count = copy.deepcopy(category_count)
    second_category_count = copy.deepcopy(category_count)
    third_category_count = copy.deepcopy(category_count)
    fourth_category_count= copy.deepcopy(category_count)
    fifth_category_count = copy.deepcopy(category_count)
    os.chdir(root_loc)
    xls = pd.ExcelFile("LVPEI_Refractive_Error_Data_Colorized_Summarized.xlsx")
    worksheets = xls.sheet_names
    #worksheets = worksheets.remove(u'Patient Visit Count')
    worksheets.pop()
    for worksheet in worksheets:
        uid_list=[]
        print worksheet
        ws = xls.parse(worksheet)
        ws[["od_sph","od_cyl","os_sph","os_cyl"]] = ws[["od_sph","od_cyl","os_sph","os_cyl"]].convert_objects(convert_numeric=True).dropna()
        #ws = ws.dropna(subset=["od_sph","od_cyl","os_sph","os_cyl"])
        uid_list = list(pd.unique(ws.uid.ravel()))
        #print len(uid_list)
        last_uid = 0
        for index, row in ws.iterrows():
            if row['uid'] in uid_list:
                first_category_count = classify(first_category_count,row)
                uid_list.remove(row['uid'])
            else:
                continue

    for worksheet in worksheets:
        ws = xls.parse(worksheet)
        ws[["od_sph","od_cyl","os_sph","os_cyl"]] = ws[["od_sph","od_cyl","os_sph","os_cyl"]].convert_objects(convert_numeric=True).dropna()
        #ws = ws.dropna(subset=["od_sph","od_cyl","os_sph","os_cyl"])
        uid_list = list(pd.unique(ws.uid.ravel()))
        for x in uid_list:
            rows = ws.loc[ws['uid']==x]
            
            first_date = 0
            zero_flag = False
            second_flag = False
            third_flag = False
            fourth_flag = False
            fifth_flag = False
            for index, row in rows.iterrows():
                if first_date==0:
                    first_date = row['visit_date']
                else:
                    if (row['visit_date'] - first_date).days >=730 and not fifth_flag:
                        fifth_category_count= classify(fifth_category_count,row)
                        fifth_flag = True
                    elif (row['visit_date'] - first_date).days >=365 and (row['visit_date'] - first_date).days < 730and not fourth_flag:
                        fourth_category_count=classify(fourth_category_count,row)
                        fourth_flag = True
                    elif (row['visit_date'] - first_date).days >=240 and (row['visit_date'] - first_date).days <365 and not third_flag:
                        third_category_count = classify(third_category_count,row)
                        third_flag = True
                    elif (row['visit_date'] - first_date).days >=120 and (row['visit_date'] - first_date).days <240 and not second_flag:
                        second_category_count=classify(second_category_count,row)
                        second_flag= True
                    elif (row['visit_date'] - first_date).days >0 and (row['visit_date'] - first_date).days <120 and not zero_flag:
                        zero_category_count=classify(zero_category_count,row)
                        zero_flag= True


    return first_category_count,zero_category_count,second_category_count,third_category_count,fourth_category_count,fifth_category_count
    
def classify(category_count,row):
    #MERGE BOTH OS OD TO SINGLE ONE, BY ADDING A PARAMETER OS/OD
    #CONVERT TO CASE SWITCH LATER
    if not np.isnan(row['os_sph']) and not np.isnan(row['os_cyl']): 
    #OS Classification
        if row['os_sph'] < 0.00:
            if row['os_cyl'] ==0.00:
                if row['os_sph'] >= -3.0 :
                    category_count['simple_myopia_low']['os'][int(row['age'])] +=1
                elif row['os_sph'] >= -6.0 and row['os_sph'] < -3.0 :
                    category_count['simple_myopia_moderate']['os'][int(row['age'])] +=1
                elif row['os_sph'] < -6.0:
                    category_count['simple_myopia_high']['os'][int(row['age'])] +=1
            elif row['os_cyl'] < 0.00:
                category_count['compound_myopic_astigmatism']['os'][int(row['age'])] +=1
            elif row['os_cyl'] >0.00:
                if abs(row['os_cyl']) > abs(row['os_sph']):
                    category_count['mixed_astigmatism_negative_sph']['os'][int(row['age'])] +=1
                else:
                    category_count['undefined']['os'][int(row['age'])] += 1
        elif row['os_sph'] == 0.00 :
            if row['os_cyl'] == 0.00:
                category_count['normal']['os'][int(row['age'])] += 1
            elif row['os_cyl'] < 0.00:
                category_count['simple_myopic_astigmatism']['os'][int(row['age'])] += 1
            elif row['os_cyl'] > 0.00:
                category_count['simple_hypermetropic_astigmatism']['os'][int(row['age'])] += 1
        elif row['os_sph'] > 0.00 :
            if row['os_cyl'] ==0.00:
                if row['os_sph'] <= 3.00:
                    category_count['simple_hyperopia_low']['os'][int(row['age'])] += 1
                elif row['os_sph'] > 3.00 and row['os_sph'] <= 6.0:
                    category_count['simple_hyperopia_moderate']['os'][int(row['age'])] += 1
                elif row['os_sph'] > 6.00 :
                    category_count['simple_hyperopia_high']['os'][int(row['age'])] += 1
            elif row['os_cyl'] > 0.00:
                category_count['compound_hypermetropic_astigmatism']['os'][int(row['age'])] +=1
            elif row['os_cyl'] < 0.00:
                if abs(row['os_sph']) > abs(row['os_cyl']):
                    category_count['mixed_astigmatism_postive_sph']['os'][int(row['age'])] +=1
                else:
                    category_count['undefined']['os'][int(row['age'])] +=1
    else:
        category_count['Insufficient Data']['os'][int(row['age'])] +=1
    
    if not np.isnan(row['od_sph']) and not np.isnan(row['od_cyl']):
        # OD Classification
        if row['od_sph'] < 0.00:
            if row['od_cyl'] ==0.00:
                if row['od_sph'] >= -3.0 :
                    category_count['simple_myopia_low']['od'][int(row['age'])] +=1
                elif row['od_sph'] >= -6.0 and row['od_sph'] < -3.0 :
                    category_count['simple_myopia_moderate']['od'][int(row['age'])] +=1
                elif row['od_sph'] < -6.0:
                    category_count['simple_myopia_high']['od'][int(row['age'])] +=1
            elif row['od_cyl'] < 0.00:
                category_count['compound_myopic_astigmatism']['od'][int(row['age'])] +=1
            elif row['od_cyl'] >0.00:
                if abs(row['od_cyl']) > abs(row['od_sph']):
                    category_count['mixed_astigmatism_negative_sph']['od'][int(row['age'])] +=1
                else:
                    category_count['undefined']['od'][int(row['age'])] += 1
        elif row['od_sph'] == 0.00 :
            if row['od_cyl'] == 0.00:
                category_count['normal']['od'][int(row['age'])] += 1
            elif row['od_cyl'] < 0.00:
                category_count['simple_myopic_astigmatism']['od'][int(row['age'])] += 1
            elif row['od_cyl'] > 0.00:
                category_count['simple_hypermetropic_astigmatism']['od'][int(row['age'])] += 1
        elif row['od_sph'] > 0.00 :
            if row['od_cyl'] ==0.00:
                if row['od_sph'] <= 3.00:
                    category_count['simple_hyperopia_low']['od'][int(row['age'])] += 1
                elif row['od_sph'] > 3.00 and row['od_sph'] <= 6.0:
                    category_count['simple_hyperopia_moderate']['od'][int(row['age'])] += 1
                elif row['od_sph'] > 6.00 :
                    category_count['simple_hyperopia_high']['od'][int(row['age'])] += 1
            elif row['od_cyl'] > 0.00:
                category_count['compound_hypermetropic_astigmatism']['od'][int(row['age'])] +=1
            elif row['od_cyl'] < 0.00:
                if abs(row['od_sph']) > abs(row['od_cyl']):
                    category_count['mixed_astigmatism_postive_sph']['od'][int(row['age'])] +=1
                else:
                    category_count['undefined']['od'][int(row['age'])] +=1
    else:
        category_count['Insufficient Data']['od'][int(row['age'])] +=1    
    return category_count
            
def analysis_write(first_category_count,zero_category_count,second_category_count,third_category_count,fourth_category_count,fifth_category_count):
    workbook = load_workbook(filename = "LVPEI_Refractive_Error_Data_Colorized_Summarized.xlsx")
    new_ws = workbook.create_sheet()
    new_ws.title = "Categorical analysis"
    rownum = 1
    colnum = 1
    Label = "First Visit"
    new_ws, rownum = format_write(new_ws,first_category_count,Label,rownum)
    Label = "4 Month Visit"
    new_ws, rownum = format_write(new_ws,second_category_count,Label,rownum)
    Label = "Before 4 Month Visit"
    new_ws, rownum = format_write(new_ws,zero_category_count,Label,rownum)
    Label = "8 Month Visit"
    new_ws, rownum = format_write(new_ws,third_category_count,Label,rownum)
    Label = "1 year Visit"
    new_ws, rownum = format_write(new_ws,fourth_category_count,Label,rownum)
    Label = "2 year Visit"
    new_ws, rownum = format_write(new_ws,fifth_category_count,Label,rownum)

    workbook.save("LVPEI_Refractive_Error_Data_Colorized_Summarized_Analyzed.xlsx")

def format_write(new_ws,first_category_count,Label,rownum,colnum=1):
    new_ws.cell(row=rownum,column=colnum).value = Label
    #Header Columns
    for i in xrange(0,26):
        colnum+=1
        new_ws.cell(row=rownum,column=colnum).value = "Age "+str(i)
        colnum+=1
        new_ws.cell(row=rownum,column=colnum).value = "Age "+str(i)
    colnum = 1
    rownum+=1
    for i in xrange(0,26):
        colnum+=1
        new_ws.cell(row=rownum,column=colnum).value = "OD"
        colnum+=1
        new_ws.cell(row=rownum,column=colnum).value = "OS"
    
    #Header Rows along with data
    rownum +=1
    
    for category in sorted(first_category_count.iterkeys()):
        colnum = 1    
        new_ws.cell(row=rownum,column=colnum).value = category
        colnum += 1
        for age,num in first_category_count[category]['od'].iteritems():
            new_ws.cell(row=rownum,column=colnum).value = num 
            colnum+=2
        colnum = 3
        for age,num in first_category_count[category]['os'].iteritems():
            new_ws.cell(row=rownum,column=colnum).value = num 
            colnum+=2
        rownum+=1
    rownum+=2
    return new_ws,rownum
                

def main():
    #age_count_dict = age_count_summary()
    #age_summary_write(age_count_dict)
    first_category_count,zero_category_count,second_category_count,third_category_count,fourth_category_count,fifth_category_count = analysis()
    #print first_category_count
    analysis_write(first_category_count,zero_category_count,second_category_count,third_category_count,fourth_category_count,fifth_category_count)

if __name__ == '__main__':
    main()