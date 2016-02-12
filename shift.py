from openpyxl import Workbook
from openpyxl import load_workbook
import os

root_loc = "/home/varunwachaspati/LVPEI/"
conditions = {"Simple Myopia Low":{'os':2,'od':2},"Simple Myopia Moderate":{'os':2,'od':2},"Simple Myopia High":{'os':2,'od':2},'Compound Myopic Astigmatism':{'os':2,'od':2},'Mixed Astigmatism (-ve Sph)':{'os':2,'od':2},'Undefined':{'os':2,'od':2},'Normal':{'os':2,'od':2},'Simple Myopic Astigmatism':{'os':2,'od':2},'Simple Hyperopia Astigmatism':{'os':2,'od':2},'Simple Hyperopia Low':{'os':2,'od':2},'Simple Hyperopia Moderate':{'os':2,'od':2},'Simple Hyperopia High':{'os':2,'od':2},'Compound Hyperopia Astigmatism':{'os':2,'od':2},'Mixed Astigmatism (+ve Sph)':{'os':2,'od':2},'Insufficient Data':{'os':2,'od':2}}

def main():
    os.chdir(root_loc)
    workbook = load_workbook(filename = "LVPEI_Refractive_Error_Data.xlsx")
    total_od_wb = Workbook()
    total_os_wb = Workbook()
    
    for x,y in conditions.iteritems():
        ws1 = total_od_wb.create_sheet()
        ws1.title = x
        ws2 = total_os_wb.create_sheet()
        ws2.title = x
        i = 1
        for ind in workbook.worksheets[0].rows[0]:
            ws1.cell(row=1 ,column=i).value = ind.value
            ws2.cell(row=1 ,column=i).value = ind.value
            i+=1
    total_od_wb.remove_sheet(total_od_wb.get_sheet_by_name('Sheet'))
    total_os_wb.remove_sheet(total_os_wb.get_sheet_by_name('Sheet'))
    for ws in workbook.worksheets:
        rows = ws.rows[1:]
        for row in rows:
            os_category = row[-1].value
            od_category = row[-2].value
            os_sheet = total_os_wb.get_sheet_by_name(os_category)
            od_sheet = total_od_wb.get_sheet_by_name(od_category)
            i=1
            for cell in row:
                if i != 1:
                    os_sheet.cell(row=conditions[os_category]['os'],column=i).value = cell.value
                    od_sheet.cell(row=conditions[od_category]['od'],column=i).value = cell.value
                else:
                    os_sheet.cell(row=conditions[os_category]['os'],column=i).value = str(ws.title[:ws.title.index('_')+1]) + str(cell.value)
                    od_sheet.cell(row=conditions[od_category]['od'],column=i).value = str(ws.title[:ws.title.index('_')+1]) + str(cell.value)
                i+=1
            conditions[os_category]['os']+=1
            conditions[od_category]['od']+=1
    total_od_wb.save('OD Categorical Analysis of Patients.xlsx')
    total_os_wb.save('OS Categorical Analysis of Patients.xlsx')


if __name__ == '__main__':
    main()
