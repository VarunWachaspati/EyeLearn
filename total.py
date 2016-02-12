import os
from openpyxl.cell import Cell
from openpyxl import load_workbook

root_loc = "/home/varunwachaspati/LVPEI/"

def categorize(sph,cyl):
    if type(sph)!=type(None) and type(cyl)!=type(None): 
        marker = "`~/'"
        sph = str(sph).replace(",",".")
        cyl = str(cyl).replace(",",".")
        for i in range(0,len(marker)):
            sph = sph.replace(marker[i],"")
            cyl = cyl.replace(marker[i],"")
        if sph.count(".") > 1:
                a=""
                flag = True
                for i in range(0,len(cyl)):
                    if sph[i]!="." or flag:
                        a+=sph[i]
                        if sph[i]==".":
                            flag = False
                sph = float(a)
        try:
            sph = float(sph)
        except:
            if not contains_number(sph):
                return "Insufficient Data"
        if not contains_number(cyl):
            return "Insufficient Data"
        
        try:
            cyl = float(cyl)
        except:
            if cyl.count(".") > 1:
                a=""
                flag = True
                for i in range(0,len(cyl)):
                    if cyl[i]!="." or flag:
                        a+=cyl[i]
                        if cyl[i]==".":
                            flag = False
                cyl = float(a)

        if sph < 0.00:
            if cyl ==0.00:
                if sph >= -3.0 :
                    return "Simple Myopia Low"
                elif sph >= -6.0 and sph < -3.0 :
                    return "Simple Myopia Moderate"
                elif sph < -6.0:
                    return "Simple Myopia High"
            elif cyl < 0.00:
                return 'Compound Myopic Astigmatism'
            elif cyl >0.00:
                if abs(cyl) > abs(sph):
                    return 'Mixed Astigmatism (-ve Sph)'
                else:
                    return 'Undefined'
        elif sph == 0.00 :
            if cyl == 0.00:
                return 'Normal'
            elif cyl < 0.00:
                return 'Simple Myopic Astigmatism'
            elif cyl > 0.00:
                return 'Simple Hyperopia Astigmatism'
        elif sph > 0.00 :
            if cyl ==0.00:
                if sph <= 3.00:
                    return 'Simple Hyperopia Low' 
                elif sph > 3.00 and sph <= 6.0:
                    return 'Simple Hyperopia Moderate'
                elif sph > 6.00 :
                    return 'Simple Hyperopia High'
            elif cyl > 0.00:
                return 'Compound Hyperopia Astigmatism'
            elif cyl < 0.00:
                if abs(sph) > abs(cyl):
                    return 'Mixed Astigmatism (+ve Sph)'
                else:
                    return 'Undefined'
    else:
        return 'Insufficient Data'

def contains_number(input_string):
    return any(char.isdigit() for char in input_string)

def category_write(row):
    od_sph = row[3].value
    od_cyl = row[4].value
    os_sph = row[8].value
    os_cyl = row[9].value
    od_category = categorize(od_sph,od_cyl)
    os_category = categorize(os_sph,os_cyl)
    return od_category, os_category

def main():
    os.chdir(root_loc)
    workbook = load_workbook(filename = "LVPEI_Refractive_Error_Data.xlsx")
    for ws in workbook.worksheets:
        ws['N1'] = "od_category"
        ws['O1'] = "os_category"
        line = 2
        rows = list(ws.rows)[1:]
        for row in rows:
            od_category, os_category = category_write(row)
            ws['N'+str(line)] = od_category
            ws['O'+str(line)] = os_category
            line +=1

    workbook.save("LVPEI_Refractive_Error_Data.xlsx")

if __name__ == '__main__':
    main()