import os
import sys
from openpyxl.cell import Cell
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill
from openpyxl import load_workbook
import openpyxl

root_loc = "/home/varunwachaspati/LVPEI/"
# AARRGGBB - Color Format
blueFill = PatternFill(start_color='774488BB', end_color='774488BB', fill_type='solid')
yellowFill = PatternFill(start_color='99FFFF00', end_color='99FFFF00', fill_type='solid')
def main():
    os.chdir(root_loc)
    workbook = load_workbook(filename = sys.argv[1])
    for ws in workbook.worksheets:
        prev_value = ws.rows[1][0].value
        prev_color = blueFill
#        print prev_value, prev_color
        row_list = ws.rows[2:]
        for cell in ws.rows[1]:
            cell.fill = blueFill
        for row in row_list:
            if row[0].value!=prev_value and prev_color==blueFill:
                for cell in row:
                    cell.fill = yellowFill
                prev_color = yellowFill
                prev_value = row[0].value
#                print "1", prev_value, prev_color
#                sys.exit(0)
            elif row[0].value!=prev_value and prev_color!=blueFill:
                for cell in row:
                    cell.fill = blueFill
                prev_color = blueFill
                prev_value = row[0].value
 #               print "2",prev_value, prev_color
 #              sys.exit(0)
            elif row[0].value==prev_value and prev_color==blueFill:
                for cell in row:
                    cell.fill = blueFill
                prev_color = blueFill
                prev_value = row[0].value
  #              print "3", prev_value, prev_color
 #               sys.exit(0)
            elif row[0].value==prev_value and prev_color!=blueFill:
                for cell in row:
                    cell.fill = yellowFill
                prev_color = yellowFill
                prev_value = row[0].value
    #            print "4", prev_value, prev_color
   #             sys.exit(0)
    workbook.save(sys.argv[1])


if __name__ == '__main__':
    main()